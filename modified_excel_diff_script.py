import argparse
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


def apply_accounting_format(worksheet):
    """Apply accounting format to all cells in a worksheet, except the header and index."""
    accounting_format = '"$"#,##0.00_);[Red]("$"#,##0.00)'
    for row in worksheet.iter_rows(min_row=2, min_col=2):  # Skip header and index
        for cell in row:
            cell.number_format = accounting_format


def autosize_columns_from_dataframe(df, worksheet):
    """Auto-size columns in an openpyxl worksheet based on dataframe contents."""
    for idx, col in enumerate(dataframe_to_rows(df, index=True, header=True), 1):
        for jdx, cell in enumerate(col, 1):
            width = max(worksheet.column_dimensions[worksheet.cell(row=idx, column=jdx).column_letter].width or 0,
                        len(str(cell)))
            worksheet.column_dimensions[worksheet.cell(row=idx, column=jdx).column_letter].width = width


def format_worksheet(worksheet):
    """Apply formatting to the worksheet."""
    # Center-align all cells
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center")
    
    # Bold the header (month) row and the "Categories" column
    for cell in worksheet["A"] + worksheet[1]:
        cell.font = Font(bold=True)


def dataframe_to_excel_sheet(df, workbook, sheet_name):
    """Write a dataframe to an Excel sheet and return the worksheet."""
    ws = workbook.create_sheet(title=sheet_name)
    for r_idx, row in enumerate(dataframe_to_rows(df, index=True, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    return ws


def enhanced_compute_differences(file1, file2, sheet_name="Monthly"):
    # Read the sheets with the specified sheet_name as the index
    monthly1 = pd.read_excel(file1, sheet_name=sheet_name, index_col='Categories')
    monthly2 = pd.read_excel(file2, sheet_name=sheet_name, index_col='Categories')

    # Outer join the two sheets for the differences
    combined = monthly1.join(monthly2, how='outer', lsuffix='_File1', rsuffix='_File2').fillna(0)
    for month in monthly1.columns:
        if month != "Yearly":  # We'll handle the "Yearly" column separately
            combined[f'Diff_{month}'] = combined[f'{month}_File1'] - combined[f'{month}_File2']
    combined['Diff_Yearly'] = combined['Yearly_File1'] - combined['Yearly_File2']

    # Create a new workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove the default sheet

    # Add each dataframe to the workbook
    ws1 = dataframe_to_excel_sheet(monthly1, wb, "File1_Monthly")
    ws2 = dataframe_to_excel_sheet(monthly2, wb, "File2_Monthly")
    ws_diff = dataframe_to_excel_sheet(combined, wb, "Diffs")

    # Apply formatting and accounting format
    for ws, df in [(ws1, monthly1), (ws2, monthly2), (ws_diff, combined)]:
        format_worksheet(ws)
        apply_accounting_format(ws)
        autosize_columns_from_dataframe(df, ws)

    
    # Add a bar chart for the differences
    chart = BarChart()
    data = Reference(ws_diff, min_col=combined.columns.get_loc('Diff_Yearly')+2,  # +2 for Index and 1-based indexing
                     min_row=1, max_col=combined.columns.get_loc('Diff_Yearly')+2, max_row=len(combined)+1)
    categories = Reference(ws_diff, min_col=1, min_row=2, max_row=len(combined)+1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.title = "Differences"
    chart.x_axis.title = "Categories"
    chart.y_axis.title = "Difference in Value"
    ws_diff.add_chart(chart, "H5")
# Save the workbook
    output_file = sheet_name.replace(' ', '_') + '_diffs.xlsx'
    wb.save(output_file)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Compare two Excel files and write the differences.')
    parser.add_argument('file1', type=str, help='First Excel file path')
    parser.add_argument('file2', type=str, help='Second Excel file path')
    parser.add_argument('--sheet_name', type=str, default="Monthly",
                        help='Sheet name to compare. Defaults to "Monthly".')

    args = parser.parse_args()
    comps = ['Monthly', 'Ideal Monthly']
    for comp in comps:
        enhanced_compute_differences(args.file1, args.file2, comp)
