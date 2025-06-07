import argparse
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


def apply_accounting_format(worksheet):
    """Apply accounting format to all cells in a worksheet, except the header and index."""
    accounting_format = '"$"#,##0.00_);[Red]("$"#,##0.00)'
    for row in worksheet.iter_rows(min_row=2, min_col=2):  # Skip header and index
        for cell in row:
            cell.number_format = accounting_format


def autosize_columns_from_dataframe(df, worksheet):
    """Auto-size columns in an openpyxl worksheet based on dataframe contents."""
    for col in worksheet.columns:
        max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
        col_letter = col[0].column_letter
        worksheet.column_dimensions[col_letter].width = max_length + 2


def format_worksheet(worksheet):
    """Apply formatting to the worksheet."""
    # Center-align all cells
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center")

    # Bold the header (month) row and the "Categories" column
    for cell in worksheet["A"] + worksheet[1]:
        cell.font = Font(bold=True)


def dataframe_to_excel_sheet(df, workbook, sheet_name):
    """Write a dataframe to an Excel sheet and return the worksheet."""
    ws = workbook.create_sheet(title=sheet_name)
    for r_idx, row in enumerate(dataframe_to_rows(df, index=True, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    return ws


def enhanced_compute_differences(file1, file2, sheet_name="Monthly"):
    # Check if the sheet has a 'Categories' column to determine how to read it
    try:
        # Read a sample to check columns
        sample_df = pd.read_excel(file1, sheet_name=sheet_name, nrows=1, engine='openpyxl')
        has_categories = 'Categories' in sample_df.columns
    except Exception:
        has_categories = False
    
    if has_categories:
        # Read the sheets with 'Categories' as the index
        monthly1 = pd.read_excel(file1, sheet_name=sheet_name, index_col='Categories', engine='openpyxl')
        monthly2 = pd.read_excel(file2, sheet_name=sheet_name, index_col='Categories', engine='openpyxl')
        use_positional = False
    else:
        # Read the sheets without setting an index (for sheets like Q Summary)
        monthly1 = pd.read_excel(file1, sheet_name=sheet_name, engine='openpyxl')
        monthly2 = pd.read_excel(file2, sheet_name=sheet_name, engine='openpyxl')
        
        # For certain sheets, force positional comparison even if first column is unique
        force_positional_sheets = ['Q Summary', 'Balances', 'Expenses']
        
        if sheet_name in force_positional_sheets:
            use_positional = True
        elif len(monthly1.columns) > 0 and monthly1.iloc[:, 0].dtype == 'object':
            first_col = monthly1.columns[0]
            # Check if the first column values are unique (good for index)
            if monthly1[first_col].nunique() == len(monthly1[first_col].dropna()):
                monthly1 = monthly1.set_index(first_col)
                monthly2 = monthly2.set_index(first_col)
                use_positional = False
            else:
                # Use positional comparison for sheets where order matters
                use_positional = True
        else:
            use_positional = True

    # Normalize column names for consistency
    monthly1.columns = monthly1.columns.str.strip()
    monthly2.columns = monthly2.columns.str.strip()

    if use_positional:
        # For positional comparison, align by row number rather than index
        max_rows = max(len(monthly1), len(monthly2))
        
        # Create a combined dataframe showing side-by-side comparison
        combined_data = {}
        
        # Add File1 data
        for col in monthly1.columns:
            combined_data[f'{col}_File1'] = list(monthly1[col]) + [0] * (max_rows - len(monthly1))
        
        # Add File2 data  
        for col in monthly2.columns:
            if col in monthly1.columns:
                combined_data[f'{col}_File2'] = list(monthly2[col]) + [0] * (max_rows - len(monthly2))
        
        # Create the combined dataframe
        combined = pd.DataFrame(combined_data)
        
        # Calculate differences for each column
        for col in monthly1.columns:
            if col in monthly2.columns:
                file1_col = f'{col}_File1'
                file2_col = f'{col}_File2'
                try:
                    # Try numeric difference
                    val1 = pd.to_numeric(combined[file1_col], errors='coerce').fillna(0)
                    val2 = pd.to_numeric(combined[file2_col], errors='coerce').fillna(0)
                    combined[f'Diff_{col}'] = val1 - val2
                except:
                    # String comparison
                    combined[f'Diff_{col}'] = (combined[file1_col].astype(str) != combined[file2_col].astype(str)).astype(int)
        
        # Add row numbers as index
        combined.index = range(len(combined))
        combined.index.name = 'Row'
        
    else:
        # Original join-based comparison for indexed data
        combined = monthly1.join(monthly2, how='outer', lsuffix='_File1', rsuffix='_File2').fillna(0)

        # Calculate differences for each column
        for col in monthly1.columns:
            if f'{col}_File1' in combined.columns and f'{col}_File2' in combined.columns:
                try:
                    # Try numeric difference first
                    combined[f'Diff_{col}'] = pd.to_numeric(combined[f'{col}_File1'], errors='coerce') - pd.to_numeric(combined[f'{col}_File2'], errors='coerce')
                except:
                    # If numeric conversion fails, just mark as different if values don't match
                    combined[f'Diff_{col}'] = (combined[f'{col}_File1'] != combined[f'{col}_File2']).astype(int)

    # Create a new workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove the default sheet

    # Add each dataframe to the workbook
    ws1 = dataframe_to_excel_sheet(monthly1, wb, "File1_" + sheet_name.replace(' ', '_'))
    ws2 = dataframe_to_excel_sheet(monthly2, wb, "File2_" + sheet_name.replace(' ', '_'))
    ws_diff = dataframe_to_excel_sheet(combined, wb, "Diffs")

    # Apply formatting and accounting format
    for ws, df in [(ws1, monthly1), (ws2, monthly2), (ws_diff, combined)]:
        format_worksheet(ws)
        apply_accounting_format(ws)
        autosize_columns_from_dataframe(df, ws)

    # Save the workbook
    output_file = sheet_name.replace(' ', '_') + '_diffs.xlsx'
    wb.save(output_file)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Compare two Excel files and write the differences.')
    parser.add_argument('file1', type=str, help='First Excel file path')
    parser.add_argument('file2', type=str, help='Second Excel file path')
    parser.add_argument('--sheet_name', type=str, default="Monthly",
                        help='Sheet name to compare. Defaults to "Monthly".')
    parser.add_argument('--all_sheets', action='store_true',
                        help='Compare all important sheets instead of just Monthly/Ideal Monthly')

    args = parser.parse_args()
    
    if args.all_sheets:
        # Compare all important sheets
        comps = ['Monthly', 'Ideal Monthly', 'Q Summary', 'Balances', 'Expenses', 'Categories']
        print("Comparing all important sheets...")
    elif args.sheet_name != "Monthly":
        # Single sheet specified
        comps = [args.sheet_name]
        print(f"Comparing {args.sheet_name} sheet...")
    else:
        # Original behavior - just Monthly and Ideal Monthly
        comps = ['Monthly', 'Ideal Monthly']
        print("Comparing Monthly and Ideal Monthly sheets...")
    
    files_compared = []
    for comp in comps:
        try:
            enhanced_compute_differences(args.file1, args.file2, comp)
            files_compared.append(comp)
            print(f"✓ Created {comp.replace(' ', '_')}_diffs.xlsx")
        except Exception as e:
            print(f"✗ Error comparing {comp}: {e}")
    
    print(f"\nComparison complete. Generated diff files for: {', '.join(files_compared)}")
